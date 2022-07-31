import csv
from openpyxl import Workbook, load_workbook
libro = Workbook()
hoja=libro.active


contenido_archivo_txt=[]
archivo_txt=open("infoBat2.txt","r")
contenido_archivo_txt=archivo_txt.readlines()

linea_numero=0
linea=""

for linea_numero in range (0, len(contenido_archivo_txt)-1):
    linea=contenido_archivo_txt[linea_numero]
    linea=linea.replace(".",";")
    contenido_archivo_txt[linea_numero]=linea
    
with open("infoBat2.csv","w") as archivo_csv:
    
    linea_numero=0
    for linea_numero in range (0, len(contenido_archivo_txt)-1):
        archivo_csv.write(contenido_archivo_txt[linea_numero])
        linea_numero = linea_numero + 1


#libro=load_workbook('infoBat2.csv')

with open('infoBat2.csv') as archivo:
    reader = csv.reader(archivo, delimiter=';')
    for row in reader:
        hoja.append(row)

cabecera = hoja.append(["Fecha", "Hora", "V_3.3", "Corriente","V_Bat","x","V_Panel"])

ultima_fila=hoja.max_row
cabecera_nombre="A"+str(ultima_fila)+":G"+str(ultima_fila)
print(cabecera_nombre)
hoja.move_range(cabecera_nombre, rows=-ultima_fila+1, cols=0)

libro.save('infoBat2.xlsx')
